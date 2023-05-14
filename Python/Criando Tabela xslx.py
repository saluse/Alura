import openpyxl

#Crinado uma planilha
book = openpyxl.Workbook()
# Como visualizar páginas existentes
print(book.sheetnames)
# Como criar uma página
book.create_sheet('Frutas')
#Como selecionar uma páginas 
frutas_page = book['Frutas']
frutas_page.append(['Nome', 'Quantidade', 'Preço'])
frutas_page.append(['Banana', '7', '3'])
frutas_page.append(['Maçã', '5', '2'])
frutas_page.append(['Uva', '3', '1'])
frutas_page.append(['Laranja', '6', '2'])
#Salvar planilha
book.save('Planilha de compras.xlsx')
print(book)